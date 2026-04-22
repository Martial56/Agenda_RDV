from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.http import JsonResponse
from datetime import date, timedelta
from .models import Patient, RendezVous, get_patients_queryset, get_rdv_queryset
from .forms import (PatientForm, RendezVousAvecPatientForm, RendezVousSansPatientForm,
                    RelancePreventiveForm, RelanceAbsenceForm, PresenceForm, DateFilterForm)


def _prochain_numero(date_rdv):
    last = RendezVous.objects.filter(date_rdv=date_rdv).order_by('-numero_ordre').first()
    return (last.numero_ordre + 1) if last else 1


# ─── DASHBOARD ───────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    today = date.today()
    rdv_base = get_rdv_queryset(request.user)

    rdv_aujourd_hui = rdv_base.filter(date_rdv=today).select_related('patient')
    rdv_absent = rdv_base.filter(est_venu=False).select_related('patient')

    # Relances préventives J-2
    j2 = today + timedelta(days=2)
    relances_preventives = rdv_base.filter(
        date_rdv=j2,
        est_venu__isnull=True,
        date_relance_preventive__isnull=True,
    ).select_related('patient')

    # Relances absence J+7
    max_date = today - timedelta(days=1)
    min_date = today - timedelta(days=7)
    relances_absence = rdv_base.filter(
        date_rdv__range=(min_date, max_date),
        est_venu=False,
        date_relance_absence__isnull=True,
    ).select_related('patient')
    
    # PDV de plus rupture : RDV absents depuis plus de 28 jours
    seuil = date.today() - timedelta(days=28)

    pdv = rdv_base.filter(
        date_rdv__lt=seuil,
        est_venu=False
    ).select_related('patient')

    agenda_7j = [
        {'date': today + timedelta(days=i),
         'count': rdv_base.filter(date_rdv=today + timedelta(days=i)).count()}
        for i in range(7)
    ]

    return render(request, 'agenda/dashboard.html', {
        'today': today,
        'rdv_aujourd_hui': rdv_aujourd_hui,
        'rdv_venus':    rdv_aujourd_hui.filter(est_venu=True).count(),
        'rdv_absents': rdv_absent.count(),
        'rdv_en_attente': rdv_aujourd_hui.filter(est_venu__isnull=True).count(),
        'relances_preventives':    relances_preventives,
        'relances_absence':        relances_absence,
        'nb_relances_preventives': relances_preventives.count(),
        'nb_relances_absence':     relances_absence.count(),
        'total_patients': get_patients_queryset(request.user).count(),
        'pdv': pdv,
        'nb_pdv': pdv.count(),
        'rdv_ce_mois': rdv_base.filter(
            date_rdv__month=today.month, date_rdv__year=today.year).count(),
        'agenda_7j': agenda_7j,
    })


# ─── AGENDA ──────────────────────────────────────────────────────────────────

@login_required
def agenda_jour(request):
    form = DateFilterForm(request.GET)
    selected_date = date.today()
    if form.is_valid() and form.cleaned_data.get('date'):
        selected_date = form.cleaned_data['date']

    rdv_list = get_rdv_queryset(request.user).filter(
        date_rdv=selected_date).select_related('patient').order_by('numero_ordre')

    return render(request, 'agenda/agenda_jour.html', {
        'rdv_list': rdv_list,
        'selected_date': selected_date,
        'form': form,
        'prev_date': (selected_date - timedelta(days=1)).strftime('%Y-%m-%d'),
        'next_date': (selected_date + timedelta(days=1)).strftime('%Y-%m-%d'),
        'stats': {
            'total':      rdv_list.count(),
            'venus':      rdv_list.filter(est_venu=True).count(),
            'absents':    rdv_list.filter(est_venu=False).count(),
            'en_attente': rdv_list.filter(est_venu__isnull=True).count(),
        },
    })


# ─── PATIENTS ────────────────────────────────────────────────────────────────

@login_required
def patient_list(request):
    query = request.GET.get('q', '')
    patients = get_patients_queryset(request.user).order_by('-date_creation')
    if query:
        patients = patients.filter(
            Q(numero_unique__icontains=query) |
            Q(residence_adresse__icontains=query) |
            Q(contact_telephone__icontains=query)
        )
    return render(request, 'agenda/patient_list.html', {'patients': patients, 'query': query})


@login_required
def patient_create(request):
    if request.method == 'POST':
        form = PatientForm(request.POST)
        if form.is_valid():
            patient = form.save(commit=False)
            patient.cree_par = request.user
            patient.save()
            messages.success(request, f"Patient {patient.numero_unique} créé avec succès.")
            return redirect('rdv_create_for_patient', patient_id=patient.pk)
    else:
        form = PatientForm()
    return render(request, 'agenda/patient_form.html', {'form': form, 'title': 'Nouveau Patient'})


@login_required
def patient_detail(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    rdv_list = patient.rendez_vous.all().order_by('-date_rdv')
    prochain_rdv = patient.rendez_vous.filter(
        date_rdv__gte=date.today(), est_venu__isnull=True
    ).order_by('date_rdv').first()
    return render(request, 'agenda/patient_detail.html', {
        'patient': patient, 'rdv_list': rdv_list, 'prochain_rdv': prochain_rdv,
    })


@login_required
def patient_edit(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    if request.method == 'POST':
        form = PatientForm(request.POST, instance=patient)
        if form.is_valid():
            form.save()
            messages.success(request, "Patient mis à jour.")
            return redirect('patient_detail', pk=pk)
    else:
        form = PatientForm(instance=patient)
    return render(request, 'agenda/patient_form.html', {
        'form': form, 'title': 'Modifier Patient', 'patient': patient})


# ─── RENDEZ-VOUS ─────────────────────────────────────────────────────────────

@login_required
def rdv_create(request):
    """Depuis la sidebar : l'utilisateur choisit le patient dans une liste déroulante."""
    if request.method == 'POST':
        form = RendezVousAvecPatientForm(request.POST)
        if form.is_valid():
            rdv = form.save(commit=False)
            rdv.cree_par = request.user
            rdv.numero_ordre = _prochain_numero(rdv.date_rdv)
            rdv.save()
            messages.success(request,
                f"RDV du {rdv.date_rdv.strftime('%d/%m/%Y')} enregistré pour {rdv.patient.numero_unique}.")
            return redirect('agenda_jour')
        # form invalide → réafficher avec erreurs (pas de redirect)
    else:
        initial = {}
        if request.GET.get('date'):
            initial['date_rdv'] = request.GET.get('date')
        form = RendezVousAvecPatientForm(initial=initial)

    return render(request, 'agenda/rdv_form.html', {
        'form': form,
        'title': 'Planifier un RDV',
        'patient': None,   # pas de patient pré-sélectionné
    })


@login_required
def rdv_create_for_patient(request, patient_id):
    """Depuis la fiche patient : patient déjà connu, formulaire sans champ patient."""
    patient = get_object_or_404(Patient, pk=patient_id)
    if request.method == 'POST':
        form = RendezVousSansPatientForm(request.POST)
        if form.is_valid():
            rdv = form.save(commit=False)
            rdv.patient = patient          # ← patient assigné ici, pas dans le formulaire
            rdv.cree_par = request.user
            rdv.numero_ordre = _prochain_numero(rdv.date_rdv)
            rdv.save()
            messages.success(request,
                f"RDV du {rdv.date_rdv.strftime('%d/%m/%Y')} ajouté à l'agenda.")
            return redirect('patient_detail', pk=patient_id)
        # form invalide → réafficher avec erreurs
    else:
        form = RendezVousSansPatientForm()

    return render(request, 'agenda/rdv_form.html', {
        'form': form,
        'title': 'Planifier un RDV',
        'patient': patient,   # affiché dans le bandeau info, pas dans le form
    })


# ─── RELANCES ────────────────────────────────────────────────────────────────

@login_required
def a_relancer_list(request):
    today = date.today()
    j2    = today + timedelta(days=2)
    max_date = today - timedelta(days=1)
    min_date = today - timedelta(days=7)

    rdv_base = get_rdv_queryset(request.user)

    preventives = rdv_base.filter(
        date_rdv=j2,
        est_venu__isnull=True,
        date_relance_preventive__isnull=True,
    ).select_related('patient').order_by('date_rdv')

    absences = rdv_base.filter(
        date_rdv__range=(min_date, max_date),
        est_venu=False,
        date_relance_absence__isnull=True,
    ).select_related('patient').order_by('date_rdv')
    
    seuil = date.today() - timedelta(days=28)

    pdv = rdv_base.filter(
        date_rdv__lt=seuil,
        est_venu=False
    ).select_related('patient').order_by('date_rdv')
    
    for rdv in pdv:
        rdv.jours_absent = (date.today() - rdv.date_rdv).days

    return render(request, 'agenda/a_relancer.html', {
        'preventives': preventives,
        'absences':    absences,
        'pdv':           pdv,
        'today':       today,
        'j2':          j2,
    })


def _creer_rdv_reprogramme(rdv_origine, nouvelle_date, user):
    RendezVous.objects.create(
        patient=rdv_origine.patient,
        date_rdv=nouvelle_date,
        date_derniere_visite=rdv_origine.date_derniere_visite,
        motif_rdv=rdv_origine.motif_rdv,
        cree_par=user,
        numero_ordre=_prochain_numero(nouvelle_date),
    )


@login_required
def rdv_relance_preventive(request, pk):
    rdv = get_object_or_404(RendezVous, pk=pk)
    if request.method == 'POST':
        form = RelancePreventiveForm(request.POST, instance=rdv)
        if form.is_valid():
            saved = form.save()
            if saved.resultat_relance_preventive == '3' and saved.date_rdv_reprogramme:
                _creer_rdv_reprogramme(rdv, saved.date_rdv_reprogramme, request.user)
                messages.success(request,
                    f"Relance préventive enregistrée. RDV reprogrammé au "
                    f"{saved.date_rdv_reprogramme.strftime('%d/%m/%Y')}.")
            else:
                messages.success(request, "Relance préventive enregistrée.")
            return redirect('a_relancer_list')
    else:
        form = RelancePreventiveForm(instance=rdv,
                                      initial={'date_relance_preventive': date.today()})
    return render(request, 'agenda/relance_form.html', {
        'form': form, 'rdv': rdv, 'type_relance': 'preventive'})


@login_required
def rdv_relance_absence(request, pk):
    rdv = get_object_or_404(RendezVous, pk=pk)
    if request.method == 'POST':
        form = RelanceAbsenceForm(request.POST, instance=rdv)
        if form.is_valid():
            saved = form.save()
            if saved.resultat_relance_absence == '3' and saved.date_rdv_reprogramme:
                _creer_rdv_reprogramme(rdv, saved.date_rdv_reprogramme, request.user)
                messages.success(request,
                    f"Relance absence enregistrée. RDV reprogrammé au "
                    f"{saved.date_rdv_reprogramme.strftime('%d/%m/%Y')}.")
            else:
                messages.success(request, "Relance absence enregistrée.")
            return redirect('a_relancer_list')
    else:
        form = RelanceAbsenceForm(instance=rdv,
                                   initial={'date_relance_absence': date.today()})
    return render(request, 'agenda/relance_form.html', {
        'form': form, 'rdv': rdv, 'type_relance': 'absence'})


# ─── PRÉSENCE ────────────────────────────────────────────────────────────────

@login_required
def rdv_presence(request, pk):
    rdv = get_object_or_404(RendezVous, pk=pk)
    if request.method == 'POST':
        form = PresenceForm(request.POST, instance=rdv)
        if form.is_valid():
            saved = form.save()
            if saved.est_venu and saved.date_prochain_rdv:
                RendezVous.objects.create(
                    patient=rdv.patient,
                    date_rdv=saved.date_prochain_rdv,
                    date_derniere_visite=rdv.date_rdv,
                    motif_rdv='4',
                    cree_par=request.user,
                    numero_ordre=_prochain_numero(saved.date_prochain_rdv),
                )
                messages.success(request,
                    f"Présence enregistrée. Prochain RDV créé le "
                    f"{saved.date_prochain_rdv.strftime('%d/%m/%Y')}.")
            else:
                messages.success(request, "Absence enregistrée.")
            return redirect('agenda_jour')
    else:
        form = PresenceForm(instance=rdv)
    return render(request, 'agenda/presence_form.html', {'form': form, 'rdv': rdv})


# ─── API ─────────────────────────────────────────────────────────────────────

@login_required
def api_stats(request):
    today = date.today()
    rdv_base = get_rdv_queryset(request.user)
    stats = []
    for i in range(30, -1, -1):
        d = today - timedelta(days=i)
        rdv = rdv_base.filter(date_rdv=d)
        stats.append({
            'date':    d.strftime('%d/%m'),
            'total':   rdv.count(),
            'venus':   rdv.filter(est_venu=True).count(),
            'absents': rdv.filter(est_venu=False).count(),
        })
    return JsonResponse({'stats': stats})


# ─── PDV (Perdus De Vue) ─────────────────────────────────────────────────────
@login_required
def pdv_relance(request, pk):
    rdv = get_object_or_404(RendezVous, pk=pk)

    # 🔴 Vérifier si le patient est vraiment PDV (> 28 jours d'absence)
    if not rdv.date_rdv:
        messages.error(request, "Date du rendez-vous invalide.")
        return redirect('a_relancer_list')

    delta = date.today() - rdv.date_rdv

    if rdv.est_venu or delta.days <= 28:
        messages.warning(request, "Ce patient n'est pas encore considéré comme perdu de vue.")
        return redirect('a_relancer_list')

    # ✅ Traitement du formulaire
    if request.method == 'POST':
        form = RelanceAbsenceForm(request.POST, instance=rdv)

        if form.is_valid():
            saved = form.save()

            # 🔁 Reprogrammation si nécessaire
            if saved.resultat_relance_absence == '3' and saved.date_rdv_reprogramme:
                _creer_rdv_reprogramme(rdv, saved.date_rdv_reprogramme, request.user)

                messages.success(
                    request,
                    f"Patient PDV relancé. RDV reprogrammé au "
                    f"{saved.date_rdv_reprogramme.strftime('%d/%m/%Y')}."
                )
            else:
                messages.success(request, "Relance du patient PDV enregistrée.")

            return redirect('pdv_list')  # 🔥 liste spécifique PDV

    else:
        form = RelanceAbsenceForm(
            instance=rdv,
            initial={'date_relance_absence': date.today()}
        )

    return render(request, 'agenda/relance_form.html', {
        'form': form,
        'rdv': rdv,
        'type_relance': 'pdv'  # 🔥 différenciation UI
    })

@login_required
def pdv_list(request):
    seuil = date.today() - timedelta(days=28)

    pdv_rdv = get_rdv_queryset(request.user).filter(
        date_rdv__lt=seuil,
        est_venu=False
    ).select_related('patient')
    
    for rdv in pdv_rdv:
        rdv.jours_absent = (date.today() - rdv.date_rdv).days

    return render(request, 'agenda/pdv_list.html', {
        'rdv_list': pdv_rdv
    })
    
    
# API de recherche de patients pour le champ select2 du formulaire de RDV

def patient_search_rdv(request):
    term = request.GET.get('term', '')

    patients = get_patients_queryset(request.user).filter(
        numero_unique__icontains=term
    )[:10]

    results = [
        {
            'id': p.id,
            'text': f"{p.numero_unique} - {p.get_sexe_display()} - {p.age_display}"
        }
        for p in patients
    ]

    return JsonResponse({'results': results})

# ─── RAPPORTS ────────────────────────────────────────────────────────────────

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


def _style_header(ws, row, headers, fill_color="1F4E79"):
    """Applique le style sur la ligne d'en-tête."""
    fill = PatternFill("solid", start_color=fill_color, end_color=fill_color)
    font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border


def _style_row(ws, row, nb_cols, alt=False):
    """Applique le style zebra sur les lignes de données."""
    fill = PatternFill("solid", start_color="D9E1F2" if alt else "FFFFFF",
                       end_color="D9E1F2" if alt else "FFFFFF")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    font = Font(name="Arial", size=9)
    for col_idx in range(1, nb_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = fill
        cell.border = border
        cell.font = font
        cell.alignment = Alignment(vertical='center')


def _add_title(ws, title, nb_cols, date_debut, date_fin):
    """Ajoute le titre et la période en haut du tableau."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, name="Arial", size=12, color="1F4E79")
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nb_cols)
    periode = f"Période : du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
    c2 = ws.cell(row=2, column=1, value=periode)
    c2.font = Font(italic=True, name="Arial", size=10, color="595959")
    c2.alignment = Alignment(horizontal='center')

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 20


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)


def _type_client_label(code):
    from .models import TYPE_CLIENT_CHOICES
    d = dict(TYPE_CLIENT_CHOICES)
    if not code:
        return ""
    codes = code.split(',')
    return ", ".join(d.get(c.strip(), c.strip()) for c in codes)


def _motif_label(code):
    from .models import MOTIF_RDV_CHOICES
    d = dict(MOTIF_RDV_CHOICES)
    if not code:
        return ""
    codes = code.split(',')
    return ", ".join(d.get(c.strip(), c.strip()) for c in codes)


@login_required
def rapport_page(request):
    from .forms import DateFilterForm as _DFF
    from django import forms as dj_forms

    class RapportFilterForm(dj_forms.Form):
        date_debut = dj_forms.DateField(
            required=True,
            widget=dj_forms.DateInput(attrs={'class': 'form-control', 'type': 'date'}),
            label="Date début"
        )
        date_fin = dj_forms.DateField(
            required=True,
            widget=dj_forms.DateInput(attrs={'class': 'form-control', 'type': 'date'}),
            label="Date fin"
        )

    form = RapportFilterForm(request.GET or None)
    context = {'form': form}

    if form.is_valid():
        context['date_debut'] = form.cleaned_data['date_debut']
        context['date_fin'] = form.cleaned_data['date_fin']

    return render(request, 'agenda/rapport.html', context)


def _build_excel_charge_virale(rdv_qs, date_debut, date_fin):
    """1 — Patients attendus pour charge virale."""
    rdv = rdv_qs.filter(
        date_rdv__range=(date_debut, date_fin),
        motif_rdv__contains='3'
    ).select_related('patient').order_by('date_rdv', 'numero_ordre')

    wb = Workbook()
    ws = wb.active
    ws.title = "Charge Virale"

    headers = ['N°', 'N° Unique Patient', 'Sexe', 'Âge', 'Type Client',
               'Date RDV', 'N° Ordre', 'Motif RDV', 'Statut']
    nb = len(headers)
    _add_title(ws, "Liste des patients attendus pour Charge Virale", nb, date_debut, date_fin)
    _style_header(ws, 3, headers)

    for i, rdv_item in enumerate(rdv, 1):
        r = i + 3
        p = rdv_item.patient
        statut = {'en_attente': 'En attente', 'honore': 'Honoré', 'manque': 'Manqué'}.get(rdv_item.statut, '')
        ws.cell(r, 1, i)
        ws.cell(r, 2, p.numero_unique)
        ws.cell(r, 3, p.get_sexe_display())
        ws.cell(r, 4, p.age_display)
        ws.cell(r, 5, _type_client_label(p.type_client))
        ws.cell(r, 6, rdv_item.date_rdv.strftime('%d/%m/%Y'))
        ws.cell(r, 7, rdv_item.numero_ordre)
        ws.cell(r, 8, _motif_label(rdv_item.motif_rdv))
        ws.cell(r, 9, statut)
        _style_row(ws, r, nb, alt=(i % 2 == 0))

    # Total
    total_row = len(rdv) + 4
    ws.cell(total_row, 1, "TOTAL")
    ws.cell(total_row, 2, f"=COUNTA(B4:B{total_row-1})")
    ws.cell(total_row, 1).font = Font(bold=True, name="Arial")
    ws.cell(total_row, 2).font = Font(bold=True, name="Arial")

    _auto_width(ws)
    ws.freeze_panes = "A4"
    return wb


def _build_excel_patients_attendus(rdv_qs, date_debut, date_fin):
    """2 — Tous les patients attendus (RDV planifiés sur la période)."""
    rdv = rdv_qs.filter(
        date_rdv__range=(date_debut, date_fin)
    ).select_related('patient').order_by('date_rdv', 'numero_ordre')

    wb = Workbook()
    ws = wb.active
    ws.title = "Patients Attendus"

    headers = ['N°', 'N° Unique Patient', 'Sexe', 'Âge', 'Type Client',
               'Date RDV', 'N° Ordre', 'Motif RDV', 'Dernier Contact', 'Statut']
    nb = len(headers)
    _add_title(ws, "Liste des patients attendus", nb, date_debut, date_fin)
    _style_header(ws, 3, headers)

    for i, rdv_item in enumerate(rdv, 1):
        r = i + 3
        p = rdv_item.patient
        statut = {'en_attente': 'En attente', 'honore': 'Honoré', 'manque': 'Manqué'}.get(rdv_item.statut, '')
        ws.cell(r, 1, i)
        ws.cell(r, 2, p.numero_unique)
        ws.cell(r, 3, p.get_sexe_display())
        ws.cell(r, 4, p.age_display)
        ws.cell(r, 5, _type_client_label(p.type_client))
        ws.cell(r, 6, rdv_item.date_rdv.strftime('%d/%m/%Y'))
        ws.cell(r, 7, rdv_item.numero_ordre)
        ws.cell(r, 8, _motif_label(rdv_item.motif_rdv))
        ws.cell(r, 9, rdv_item.date_derniere_visite.strftime('%d/%m/%Y') if rdv_item.date_derniere_visite else '')
        ws.cell(r, 10, statut)
        _style_row(ws, r, nb, alt=(i % 2 == 0))

    total_row = len(rdv) + 4
    ws.cell(total_row, 1, "TOTAL")
    ws.cell(total_row, 2, f"=COUNTA(B4:B{total_row-1})")
    ws.cell(total_row, 1).font = Font(bold=True, name="Arial")
    ws.cell(total_row, 2).font = Font(bold=True, name="Arial")

    _auto_width(ws)
    ws.freeze_panes = "A4"
    return wb


def _build_excel_rdv_manques(rdv_qs, date_debut, date_fin):
    """3 — Patients ayant manqué leur RDV."""
    rdv = rdv_qs.filter(
        date_rdv__range=(date_debut, date_fin),
        est_venu=False
    ).select_related('patient').order_by('date_rdv', 'numero_ordre')

    wb = Workbook()
    ws = wb.active
    ws.title = "RDV Manqués"

    headers = ['N°', 'N° Unique Patient', 'Sexe', 'Âge', 'Type Client',
               'Date RDV', 'Motif RDV', 'Motif Non-venue',
               'Date Relance Absence', 'Résultat Relance']
    nb = len(headers)
    _add_title(ws, "Liste des patients ayant manqué leur RDV", nb, date_debut, date_fin)
    _style_header(ws, 3, headers, fill_color="C00000")

    from .models import RESULTAT_RELANCE_CHOICES
    res_dict = dict(RESULTAT_RELANCE_CHOICES)

    for i, rdv_item in enumerate(rdv, 1):
        r = i + 3
        p = rdv_item.patient
        ws.cell(r, 1, i)
        ws.cell(r, 2, p.numero_unique)
        ws.cell(r, 3, p.get_sexe_display())
        ws.cell(r, 4, p.age_display)
        ws.cell(r, 5, _type_client_label(p.type_client))
        ws.cell(r, 6, rdv_item.date_rdv.strftime('%d/%m/%Y'))
        ws.cell(r, 7, _motif_label(rdv_item.motif_rdv))
        ws.cell(r, 8, rdv_item.motif_non_venue or '')
        ws.cell(r, 9, rdv_item.date_relance_absence.strftime('%d/%m/%Y') if rdv_item.date_relance_absence else '')
        ws.cell(r, 10, res_dict.get(rdv_item.resultat_relance_absence, '') if rdv_item.resultat_relance_absence else '')
        _style_row(ws, r, nb, alt=(i % 2 == 0))

    total_row = len(rdv) + 4
    ws.cell(total_row, 1, "TOTAL")
    ws.cell(total_row, 2, f"=COUNTA(B4:B{total_row-1})")
    ws.cell(total_row, 1).font = Font(bold=True, name="Arial")
    ws.cell(total_row, 2).font = Font(bold=True, name="Arial")

    _auto_width(ws)
    ws.freeze_panes = "A4"
    return wb


def _build_excel_rdv_honores(rdv_qs, date_debut, date_fin):
    """4 — Patients ayant honoré leur RDV."""
    rdv = rdv_qs.filter(
        date_rdv__range=(date_debut, date_fin),
        est_venu=True
    ).select_related('patient').order_by('date_rdv', 'numero_ordre')

    wb = Workbook()
    ws = wb.active
    ws.title = "RDV Honorés"

    headers = ['N°', 'N° Unique Patient', 'Sexe', 'Âge', 'Type Client',
               'Date RDV', 'Motif RDV', 'Date Prochain RDV', 'Motif Prochain RDV']
    nb = len(headers)
    _add_title(ws, "Liste des patients ayant respecté leur RDV", nb, date_debut, date_fin)
    _style_header(ws, 3, headers, fill_color="375623")

    for i, rdv_item in enumerate(rdv, 1):
        r = i + 3
        p = rdv_item.patient
        ws.cell(r, 1, i)
        ws.cell(r, 2, p.numero_unique)
        ws.cell(r, 3, p.get_sexe_display())
        ws.cell(r, 4, p.age_display)
        ws.cell(r, 5, _type_client_label(p.type_client))
        ws.cell(r, 6, rdv_item.date_rdv.strftime('%d/%m/%Y'))
        ws.cell(r, 7, _motif_label(rdv_item.motif_rdv))
        ws.cell(r, 8, rdv_item.date_prochain_rdv.strftime('%d/%m/%Y') if rdv_item.date_prochain_rdv else '')
        ws.cell(r, 9, _motif_label(rdv_item.motif_prochain_rdv) if rdv_item.motif_prochain_rdv else '')
        _style_row(ws, r, nb, alt=(i % 2 == 0))

    total_row = len(rdv) + 4
    ws.cell(total_row, 1, "TOTAL")
    ws.cell(total_row, 2, f"=COUNTA(B4:B{total_row-1})")
    ws.cell(total_row, 1).font = Font(bold=True, name="Arial")
    ws.cell(total_row, 2).font = Font(bold=True, name="Arial")

    _auto_width(ws)
    ws.freeze_panes = "A4"
    return wb


@login_required
def rapport_export(request, rapport_type):
    from django import forms as dj_forms

    class RapportFilterForm(dj_forms.Form):
        date_debut = dj_forms.DateField(required=True)
        date_fin = dj_forms.DateField(required=True)

    form = RapportFilterForm(request.GET)
    if not form.is_valid():
        messages.error(request, "Dates invalides.")
        return redirect('rapport_page')

    date_debut = form.cleaned_data['date_debut']
    date_fin = form.cleaned_data['date_fin']
    rdv_qs = get_rdv_queryset(request.user)

    builders = {
        'charge_virale':    (_build_excel_charge_virale,    "charge_virale"),
        'patients_attendus': (_build_excel_patients_attendus, "patients_attendus"),
        'rdv_manques':      (_build_excel_rdv_manques,       "rdv_manques"),
        'rdv_honores':      (_build_excel_rdv_honores,       "rdv_honores"),
    }

    if rapport_type not in builders:
        messages.error(request, "Type de rapport inconnu.")
        return redirect('rapport_page')

    builder_fn, file_slug = builders[rapport_type]
    wb = builder_fn(rdv_qs, date_debut, date_fin)

    filename = f"{file_slug}_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ─── BACKUP ──────────────────────────────────────────────────────────────────

import zipfile
import os
from django.conf import settings
from django.contrib.auth.decorators import user_passes_test


def _dump_sqlite(db_path, now_str, zf):
    """Copie le fichier SQLite directement dans le ZIP."""
    if db_path.exists():
        zf.write(db_path, arcname=f'backup_{now_str}/database/db.sqlite3')
        # Ajoute aussi un dump SQL lisible pour portabilité
        import sqlite3
        sql_buf = io.StringIO()
        con = sqlite3.connect(str(db_path))
        for line in con.iterdump():
            sql_buf.write(line + '\n')
        con.close()
        zf.writestr(f'backup_{now_str}/database/db_dump.sql', sql_buf.getvalue())


def _dump_postgresql(db_conf, now_str, zf):
    """
    Utilise pg_dump pour exporter la base PostgreSQL en SQL.
    Fonctionne avec ou sans mot de passe (via PGPASSWORD env).
    """
    import subprocess
    import tempfile

    host     = db_conf.get('HOST', 'localhost') or 'localhost'
    port     = str(db_conf.get('PORT', 5432) or 5432)
    name     = db_conf.get('NAME', '')
    user     = db_conf.get('USER', '')
    password = db_conf.get('PASSWORD', '')

    env = os.environ.copy()
    if password:
        env['PGPASSWORD'] = password

    cmd = [
        'pg_dump',
        '-h', host,
        '-p', port,
        '-U', user,
        '--no-password',
        '--format=plain',
        '--encoding=UTF8',
        name,
    ]

    result = subprocess.run(
        cmd,
        capture_output=True,
        env=env,
        timeout=120,
    )

    if result.returncode != 0:
        err = result.stderr.decode('utf-8', errors='replace')
        raise RuntimeError(f"pg_dump a échoué : {err}")

    sql_content = result.stdout.decode('utf-8', errors='replace')
    zf.writestr(f'backup_{now_str}/database/{name}_dump.sql', sql_content)


@user_passes_test(lambda u: u.is_staff, login_url='/login/')
def backup_download(request):
    """
    Sauvegarde universelle : fonctionne avec SQLite ET PostgreSQL.
    - SQLite  → copie du fichier .sqlite3 + dump SQL texte
    - PostgreSQL → dump via pg_dump (nécessite pg_dump installé sur le serveur)
    Réservé aux utilisateurs is_staff.
    """
    from datetime import datetime
    now_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    zip_filename = f"arv_agenda_backup_{now_str}.zip"

    buf = io.BytesIO()
    base_dir = settings.BASE_DIR
    db_conf = settings.DATABASES.get('default', {})
    engine  = db_conf.get('ENGINE', '')

    INCLUDE_DIRS  = ['agenda', 'synthese', 'arv_agenda', 'templates', 'static']
    INCLUDE_FILES = ['manage.py', 'requirements.txt', 'start.sh']
    EXCLUDE_IN_PATH = ('__pycache__', '.git')
    EXCLUDE_EXTS    = ('.pyc', '.pyo')

    def should_exclude(path_str):
        p = str(path_str)
        return (any(e in p for e in EXCLUDE_IN_PATH) or
                any(p.endswith(ext) for ext in EXCLUDE_EXTS))

    db_error = None

    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:

        # ── 1. Sauvegarde de la base de données ──────────────────────────────
        try:
            if 'sqlite3' in engine:
                db_path = settings.BASE_DIR / 'db.sqlite3'
                _dump_sqlite(db_path, now_str, zf)
            elif 'postgresql' in engine or 'postgis' in engine:
                _dump_postgresql(db_conf, now_str, zf)
            else:
                zf.writestr(
                    f'backup_{now_str}/database/README.txt',
                    f"Moteur de base de données non supporté automatiquement : {engine}\n"
                    "Veuillez effectuer la sauvegarde manuellement."
                )
        except Exception as exc:
            db_error = str(exc)
            zf.writestr(
                f'backup_{now_str}/database/ERREUR.txt',
                f"Erreur lors de la sauvegarde de la base :\n{db_error}"
            )

        # ── 2. Fichiers racine ────────────────────────────────────────────────
        for fname in INCLUDE_FILES:
            fpath = base_dir / fname
            if fpath.exists():
                zf.write(fpath, arcname=f'backup_{now_str}/{fname}')

        # ── 3. Dossiers applicatifs ───────────────────────────────────────────
        for dirname in INCLUDE_DIRS:
            dir_path = base_dir / dirname
            if not dir_path.exists():
                continue
            for root, dirs, files in os.walk(dir_path):
                dirs[:] = [d for d in dirs if d not in EXCLUDE_IN_PATH]
                for file in files:
                    full_path = os.path.join(root, file)
                    if should_exclude(full_path):
                        continue
                    rel_path = os.path.relpath(full_path, base_dir)
                    zf.write(full_path, arcname=f'backup_{now_str}/{rel_path}')

        # ── 4. Fiche de synthèse ──────────────────────────────────────────────
        from datetime import datetime as dt
        info_lines = [
            f"ARV Agenda — Sauvegarde système",
            f"Date       : {dt.now().strftime('%d/%m/%Y à %H:%M:%S')}",
            f"Moteur DB  : {engine}",
            f"Statut DB  : {'✅ OK' if not db_error else '❌ ERREUR : ' + db_error}",
            f"",
            f"Contenu du backup :",
            f"  backup_{now_str}/database/  → dump de la base de données",
            f"  backup_{now_str}/agenda/    → application agenda",
            f"  backup_{now_str}/synthese/  → application synthese",
            f"  backup_{now_str}/templates/ → templates HTML",
            f"  backup_{now_str}/static/    → fichiers CSS/JS",
            f"  backup_{now_str}/manage.py  → script Django",
            f"",
            f"Restauration SQLite  : copier db.sqlite3 à la racine du projet",
            f"Restauration Postgres: psql -U <user> -d <dbname> < database/<name>_dump.sql",
        ]
        zf.writestr(f'backup_{now_str}/BACKUP_INFO.txt', '\n'.join(info_lines))

    buf.seek(0)

    if db_error:
        messages.warning(request,
            f"⚠️ Backup généré avec une erreur sur la base de données : {db_error}")

    response = HttpResponse(buf.read(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
    return response
